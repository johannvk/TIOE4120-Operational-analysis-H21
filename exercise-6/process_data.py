import openpyxl as xl
import pyomo.environ as pyo

from openpyxl.utils import coordinate_to_tuple

def last_legering_diameter_data():
    wb = xl.load_workbook("Produkt_miks.xlsx")

    basis_sheet = wb["Legering-Diameter-Data"]

    # Extract diameters and diameter percentages: 
    diameter_navn = []
    diametere = []
    for row in basis_sheet.iter_rows(min_row=2, max_row=6, min_col=1, max_col=2):
        diameter_navn.append(row[0].value)
        diametere.append(row[1].value)

    legerings_navn = []
    legeringer = []
    for row in basis_sheet.iter_rows(min_row=2, max_row=5, min_col=3, max_col=4):
        legerings_navn.append(row[0].value)
        legeringer.append(row[1].value)
        
    # Laste historiske snittverdier.
        
    historie_sheet = wb["Historie"]
    rho_cells = historie_sheet['D17':'G21']
    rho = []
    
    for row in rho_cells:
        row_values = []
        for cell in row:
            row_values.append(cell.value)
        rho.append(row_values)
    

    return {"diameter_andel_sum": diametere, "diameter_navn": diameter_navn,
            "legerings_andel_sum": legeringer, "legerings_navn": legerings_navn,
            "historisk_snitt": rho}


def skriv_løsning_til_fil(m: pyo.Model, sheet_name: str, upper_left_corner_cell: str, filename="Produkt_miks.xlsx"):
    rad, kolonne = coordinate_to_tuple(upper_left_corner_cell)

    wb = xl.load_workbook("Produkt_miks.xlsx")
    sheet = wb.get_sheet_by_name(sheet_name)

    løsning = m.X
    for i in m.Diametere:
        for j in m.Legeringer:
            sheet.cell(rad + i, kolonne + j, value = løsning[i, j].value)
    wb.save(filename=filename)
