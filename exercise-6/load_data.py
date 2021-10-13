import openpyxl as xl
import pandas as pd


def last_legering_diameter_data():
    wb = xl.load_workbook("Produkt_miks.xlsx")

    # sheet_names = wb.get_sheet_names()

    basis_sheet = wb.get_sheet_by_name("Legering-Diameter-Data")

    # diameter_names = [cell[0].value for cell in basis_sheet["A2:A6"]]
    # diameter_values = [cell[0].value for cell in basis_sheet["B2:B6"]]

    # Extract diameters and diameter percentages: 
    diameter_navn = []
    diametere = []
    for row in basis_sheet.iter_rows(min_row=2, max_row=6, min_col=1, max_col=2):
        diameter_navn.append(row[0].value)
        diametere.append(row[1].value)

    legerings_navn = []
    legeringer = []
    for row in basis_sheet.iter_rows(min_row=2, max_row=5, min_col=3, max_col=4):
        legerings_navn.append(row[0].value)
        legeringer.append(row[1].value)

    return {"diametere": diametere, "diameter_navn": diameter_navn, "legeringer": legeringer, "legerings_navn": legerings_navn}

